from flask import Flask, request, redirect, url_for, render_template, session, flash, g, send_from_directory, send_file, jsonify, abort
from datetime import datetime, date
import sqlite3
import os
import math
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from translations import translate, get_all_translations

# Decorator para verificar se o usuário está logado
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# ----------------------------
# Configurações do Flask
# ----------------------------
app = Flask(__name__)
app.secret_key = 'xcO9HrNnkfQaKZzWxF'

# --- Adicionar configuração de caminho absoluto do DB e static explicitamente ---
DATABASE_PATH = os.path.join(app.root_path, 'banco.db')
app.static_folder = os.path.join(app.root_path, 'static')
app.template_folder = os.path.join(app.root_path, 'templates')

# Criar diretórios necessários se não existirem
os.makedirs(os.path.join(app.static_folder, 'img'), exist_ok=True)
os.makedirs(os.path.join(app.static_folder, 'css'), exist_ok=True)
os.makedirs(os.path.join(app.static_folder, 'js'), exist_ok=True)

# Contexto global de tradução
@app.context_processor
def inject_translation():
    """Injeta funções de tradução em todos os templates."""
    lang = session.get('lang', 'pt')
    return {
        't': lambda key: translate(key, lang),
        'lang': lang,
        'translations': get_all_translations(lang)
    }

# ----------------------------
# Funções de Banco de Dados
# ----------------------------
def get_db():
    """Retorna a conexão com o banco de dados."""
    if 'db' not in g:
        # usar caminho absoluto evita erros quando o working dir muda
        g.db = sqlite3.connect(DATABASE_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def fechar_conexao(error):
    """Fecha a conexão com o banco de dados no final da requisição."""
    db = g.pop('db', None)
    if db is not None:
        db.commit()
        db.close()

def init_db():
    """Inicializa as tabelas do banco de dados."""
    db = get_db()
    cursor = db.cursor()
    cursor.execute("PRAGMA foreign_keys = ON;")
    cursor.execute('''CREATE TABLE IF NOT EXISTS estoque(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codigo_de_barras TEXT NOT NULL,
        lote TEXT NOT NULL,
        validade_int INTEGER NOT NULL,
        validade_text TEXT NOT NULL,
        produto_nome TEXT NOT NULL,
        quantidade INTEGER,
        image_path TEXT,
        categoria INTEGER NOT NULL
         )''') # Corrigi codigo_de_barras para TEXT para evitar erros com zeros a esquerda
    

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS movimentacao (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER NOT NULL,
            product_barcode TEXT NOT NULL,
            name TEXT NOT NULL,
            action TEXT NOT NULL,
            quantidade INTEGER NOT NULL,
            motivo TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (product_id) REFERENCES estoque(id)
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS responsaveis (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL UNIQUE,
            criado_em DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Adicionar responsáveis padrão se a tabela estiver vazia
    cursor.execute('SELECT COUNT(*) FROM responsaveis')
    if cursor.fetchone()[0] == 0:
        responsaveis_padrao = [
            'Equipe de Distribuição',
            'Coordenação',
            'Voluntários'
        ]
        for nome in responsaveis_padrao:
            cursor.execute('INSERT INTO responsaveis (nome) VALUES (?)', (nome,))
    
    db.commit()

def limpar_estoque_zerado():
    """
    Deleta do banco de dados todos os produtos onde a 'quantidade' é 0 ou menor.
    Isso ajuda a manter o estoque principal limpo.
    """
    db = get_db()
    cursor = db.cursor()
    
    try:
        # A instrução SQL que deleta as linhas onde a quantidade é menor ou igual a zero
        cursor.execute('DELETE FROM estoque WHERE quantidade <= 0')
        
        # O db.commit() é crucial para salvar a exclusão
        db.commit()
        
        # Opcional: Retorna o número de itens deletados para fins de log
        itens_deletados = cursor.rowcount
        print(f"Limpeza de estoque concluída. {itens_deletados} itens zerados removidos.")
        return itens_deletados
        
    except Exception as e:
        print(f"Erro ao tentar limpar o estoque zerado: {e}")
        db.rollback() # Desfaz a operação em caso de erro
        return 0
    finally:
        # Nota: g.db é fechado no teardown
        pass



# --------------------------------------------------------------------------
# FUNÇÕES AUXILIARES
# --------------------------------------------------------------------------

def validar_imagem(filename):
    """Valida se o arquivo é uma imagem permitida."""
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --------------------------------------------------------------------------
# ROTAS DE ARQUIVOS ESTÁTICOS
# --------------------------------------------------------------------------

@app.route('/static/<path:filename>')
def serve_static(filename):
    """Serve arquivos estáticos com validação."""
    file_path = os.path.join(app.static_folder, filename)
    if not os.path.isfile(file_path):
        app.logger.warning(f"Arquivo estático não encontrado: {filename}")
        abort(404, description=f"Arquivo não encontrado: {filename}")
    return send_from_directory(app.static_folder, filename)

@app.route('/static/img/<path:filename>')
def serve_image(filename):
    """Serve imagens do diretório static/img/ com validação e checagem de existência."""
    if not validar_imagem(filename):
        return "Tipo de arquivo não permitido", 400
    img_dir = os.path.join(app.static_folder, 'img')
    file_path = os.path.join(img_dir, filename)
    if not os.path.isfile(file_path):
        # Retorna 404 para facilitar identificação de links quebrados
        abort(404, description=f"Imagem não encontrada: {filename}")
    return send_from_directory(img_dir, filename)

# --------------------------------------------------------------------------
# ROTAS DA APLICAÇÃO
# --------------------------------------------------------------------------

@app.route('/', methods=['GET', 'POST'])
def login():
    """Gerencia o acesso: permite login com 'admin'/'admin' (case-insensitive)."""
    error = None
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        
        # Captura a preferência de idioma
        lang = 'es' if request.form.get('spanish') else 'pt'
        session['lang'] = lang

        if username.lower() == 'admin' and password.lower() == 'admin':
            session['user'] = username
            return redirect(url_for('usuario'))
        else:
            error = translate('login_error', lang)

    return render_template('index.html', error=error)

@app.route('/change_language/<lang>')
def change_language(lang):
    """Altera o idioma da sessão."""
    if lang in ['pt', 'es']:
        session['lang'] = lang
    # Redireciona de volta para a página anterior ou para home
    return redirect(request.referrer or url_for('home'))

@app.route('/usuario')
def usuario():
    """Página de escolha de perfil."""
    # Opcional: Adicionar verificação de session aqui
    return render_template('usuario.html')

@app.route('/creditos')
def creditos():
    """Página de créditos da equipe."""
    return render_template('creditos.html')

@app.route('/home', methods=['GET'])
def home():
    """Exibe a lista de todos os produtos em estoque."""
    from datetime import timedelta
    db = get_db()
    cursor = db.cursor()

    # Parâmetros de query (filtros e paginação)
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 24))
    categoria_filter = request.args.get('categoria')
    busca = request.args.get('q', '').strip()
    ordenar = request.args.get('ordenar', 'validade')  # Novo parâmetro de ordenação

    # Monta cláusula WHERE dinâmica
    where_clauses = []
    params = []

    if categoria_filter:
        where_clauses.append('categoria = ?')
        params.append(categoria_filter)

    if busca:
        # busca por nome (case-insensitive)
        where_clauses.append('LOWER(produto_nome) LIKE ?')
        params.append(f"%{busca.lower()}%")

    where_sql = ''
    if where_clauses:
        where_sql = 'WHERE ' + ' AND '.join(where_clauses)

    # Define a ordenação baseada no parâmetro
    order_by = 'ORDER BY validade_int ASC'  # Padrão
    if ordenar == 'nome':
        order_by = 'ORDER BY produto_nome ASC'
    elif ordenar == 'quantidade':
        order_by = 'ORDER BY quantidade DESC'
    elif ordenar == 'validade':
        order_by = 'ORDER BY validade_int ASC'

    # Conta total de items (para paginação)
    count_sql = f'SELECT COUNT(*) as cnt FROM estoque {where_sql}'
    total_rows = cursor.execute(count_sql, params).fetchone()['cnt']

    total_pages = max(1, math.ceil(total_rows / per_page))
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    # Busca os produtos com paginação
    sql = f'SELECT * FROM estoque {where_sql} {order_by} LIMIT ? OFFSET ?'
    final_params = params + [per_page, offset]
    produtos_raw = cursor.execute(sql, final_params).fetchall()
    
    # Adiciona status de validade para cada produto
    hoje_ts = int(datetime.combine(date.today(), datetime.min.time()).timestamp())
    dias_7_ts = int(datetime.combine(date.today() + timedelta(days=7), datetime.min.time()).timestamp())
    dias_15_ts = int(datetime.combine(date.today() + timedelta(days=15), datetime.min.time()).timestamp())
    
    produtos = []
    for p in produtos_raw:
        produto_dict = dict(p)
        validade_ts = produto_dict.get('validade_int', 0)
        
        # Define status de validade
        if validade_ts < hoje_ts:
            produto_dict['status_validade'] = 'vencido'
        elif validade_ts < dias_7_ts:
            produto_dict['status_validade'] = 'vence_urgente'
        elif validade_ts < dias_15_ts:
            produto_dict['status_validade'] = 'vence_proximo'
        else:
            produto_dict['status_validade'] = 'ok'
        
        produtos.append(produto_dict)

    # Estatísticas simples
    total_produtos = cursor.execute('SELECT SUM(COALESCE(quantidade,0)) as s FROM estoque').fetchone()['s'] or 0
    # itens baixos: quantidade <= 5
    produtos_baixos = cursor.execute('SELECT COUNT(*) as c FROM estoque WHERE quantidade <= 5').fetchone()['c'] or 0
    # próximo vencimento (data text mais próxima no futuro ou hoje)
    hoje_ts = int(datetime.combine(date.today(), datetime.min.time()).timestamp())
    prox = cursor.execute('SELECT validade_text FROM estoque WHERE validade_int >= ? ORDER BY validade_int ASC LIMIT 1', (hoje_ts,)).fetchone()
    vencimento_proximo = prox['validade_text'] if prox else 'Nenhum'
    # movimentações hoje
    hoje_iso = date.today().isoformat()
    movimentacoes_hoje = cursor.execute("SELECT COUNT(*) as c FROM movimentacao WHERE DATE(timestamp) = ?", (hoje_iso,)).fetchone()['c'] or 0

    # Flags de paginação
    has_prev = page > 1
    has_next = page < total_pages
    # Busca categorias distintas do estoque para popular o filtro de categorias
    try:
        cursor.execute('SELECT DISTINCT categoria FROM estoque ORDER BY categoria ASC')
        cats_rows = cursor.fetchall()
    except Exception:
        cats_rows = []

    # Traduz os rótulos das categorias usando as chaves 'category_<id>' em translations.py
    lang = session.get('lang', 'pt')
    categorias = []
    for row in cats_rows:
        # row pode ser sqlite3.Row; tenta extrair o valor
        cat_val = None
        try:
            cat_val = row['categoria']
        except Exception:
            try:
                cat_val = row[0]
            except Exception:
                cat_val = row

        try:
            cat_id = int(cat_val)
        except Exception:
            cat_id = cat_val

        label_key = f'category_{cat_id}'
        categorias.append({'id': cat_id, 'label': translate(label_key, lang)})

    # Garantir que a categoria 'Bebidas' (id 5) esteja disponível no filtro mesmo que ainda não exista no DB
    try:
        has_bebidas = any(int(c.get('id')) == 5 for c in categorias)
    except Exception:
        has_bebidas = False

    if not has_bebidas:
        categorias.append({'id': 5, 'label': translate('category_5', lang)})

    return render_template('home.html', produtos=produtos,
                           total_produtos=total_produtos,
                           produtos_baixos=produtos_baixos,
                           vencimento_proximo=vencimento_proximo,
                           movimentacoes_hoje=movimentacoes_hoje,
                           current_page=page,
                           total_pages=total_pages,
                           has_prev=has_prev,
                           has_next=has_next,
                           categorias=categorias)

# --------------------------------------------------------------------------
# ROTAS DE ENTRADA (ADICIONAR)
# --------------------------------------------------------------------------

@app.route('/adicionar_produto', methods=['GET', 'POST'])
def adicionar_produto():
    """Adiciona um novo produto ao estoque, somando a quantidade se o código de barras e a validade já existirem."""
    db = get_db()
    cursor = db.cursor()
    
    if request.method == 'POST':
        try:
            # Verifica se está em modo simplificado (apenas adicionar quantidade)
            modo_rapido = request.form.get('modo_rapido') == 'on'
            
            # Captura campos obrigatórios em qualquer modo
            codigo_de_barras = request.form['codigo_de_barras']
            validade = request.form['validade']
            quantidade_nova = int(request.form['quantidade'])
            
            # Converte a validade para timestamp (int)
            # Aceita formato brasileiro DD/MM/YYYY ou ISO YYYY-MM-DD
            try:
                # Tenta formato brasileiro primeiro: DD/MM/YYYY
                datetime_validade = datetime.strptime(validade, '%d/%m/%Y')
            except ValueError:
                # Se falhar, tenta formato ISO: YYYY-MM-DD
                datetime_validade = datetime.fromisoformat(validade)
            
            validade_int = int(datetime_validade.timestamp())
            # Garante que validade_text sempre esteja no formato brasileiro
            validade_text = datetime_validade.strftime('%d/%m/%Y')
            
            # Busca produto existente com este código de barras e validade
            cursor.execute('''
                SELECT * FROM estoque 
                WHERE codigo_de_barras = ? AND validade_text = ?
            ''', (codigo_de_barras, validade_text))
            
            produto_existente = cursor.fetchone()
            
            # Define os dados do produto baseado no modo
            if modo_rapido:
                # Modo rápido: busca dados de produto existente
                if produto_existente:
                    # Produto já existe com esta validade, usa seus dados
                    produto_nome = produto_existente['produto_nome']
                    lote = produto_existente['lote']
                    categoria = produto_existente['categoria']
                    image_path = produto_existente['image_path']
                else:
                    # Se não encontrou com a validade, tenta buscar qualquer produto com este código de barras
                    cursor.execute('''
                        SELECT * FROM estoque 
                        WHERE codigo_de_barras = ?
                        LIMIT 1
                    ''', (codigo_de_barras,))
                    
                    produto_ref = cursor.fetchone()
                    
                    if produto_ref:
                        # Usa dados do produto de referência mas com nova validade
                        produto_nome = produto_ref['produto_nome']
                        lote = produto_ref['lote']
                        categoria = produto_ref['categoria']
                        image_path = produto_ref['image_path']
                    else:
                        flash('Produto não encontrado. Use o formulário completo para adicionar um novo produto.', 'error')
                        return render_template('adicionar_estoque.html')
            else:
                # Modo normal: captura todos os campos do formulário
                lote = request.form.get('lote', '')
                produto_nome = request.form['produto_nome']
                categoria = request.form['categoria']
                image_path = request.form.get('image_path', '')
            
            # Agora processa a adição/atualização
            if produto_existente:
                # SE EXISTE: ATUALIZA (SOMA) A QUANTIDADE
                produto_id = produto_existente['id']
                quantidade_atual = produto_existente['quantidade']
                nova_quantidade_total = quantidade_atual + quantidade_nova
                
                # Atualiza a quantidade e demais metadados
                cursor.execute('''
                    UPDATE estoque
                    SET quantidade = ?, validade_int = ?, validade_text = ?, categoria = ?, produto_nome = ?, image_path = ?, lote = ?
                    WHERE id = ?
                ''', (nova_quantidade_total, validade_int, validade_text, categoria, produto_nome, image_path, lote, produto_id))
                
                # Registra movimentação de entrada pela adição de quantidade
                cursor.execute('''
                    INSERT INTO movimentacao (product_id, product_barcode, name, action, quantidade)
                    VALUES (?, ?, ?, ?, ?)
                ''', (produto_id, codigo_de_barras, produto_nome, 'entrada', quantidade_nova))

                flash(f"Quantidade do item '{produto_nome}' (Validade: {validade_text}) atualizada para {nova_quantidade_total}!", 'success')
                
            else:
                # SE NÃO EXISTE: INSERE NOVO REGISTRO
                cursor.execute('''
                    INSERT INTO estoque (codigo_de_barras, lote, validade_int, validade_text, produto_nome, quantidade, categoria, image_path)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (codigo_de_barras, lote, validade_int, validade_text, produto_nome, quantidade_nova, categoria, image_path))
                # obtém id do novo produto
                novo_id = cursor.lastrowid
                # Registra movimentação de entrada para o novo produto
                cursor.execute('''
                    INSERT INTO movimentacao (product_id, product_barcode, name, action, quantidade)
                    VALUES (?, ?, ?, ?, ?)
                ''', (novo_id, codigo_de_barras, produto_nome, 'entrada', quantidade_nova))

                flash("Produto adicionado com sucesso!", 'success')
            
            db.commit()
            return redirect(url_for('home'))
            
        except ValueError:
            error = "Erro: A quantidade deve ser um número válido."
            flash(error, 'error')
            return render_template('adicionar_estoque.html')
        except Exception as e:
            # Captura qualquer outro erro, como erro de banco de dados
            db.rollback()
            error = f"Erro ao processar os dados: {e}"
            flash(error, 'error')
            return render_template('adicionar_estoque.html')
            
    return render_template('adicionar_estoque.html')


    
# --------------------------------------------------------------------------
# ROTAS DE SAÍDA (RETIRADA)
# --------------------------------------------------------------------------

@app.route('/retirada', methods=['GET', 'POST'])
def retirada():
    """Processa a leitura do código de barras + data de validade e redireciona para a tela de quantidade."""
    db = get_db()
    cursor = db.cursor()
    
    lang = session.get('lang', 'pt')
    if request.method == 'POST':
        codigo_de_barras = request.form.get('codigo_de_barras') 
        validade = request.form.get('validade')

        if codigo_de_barras and validade:
            # Converter a data para timestamp
            try:
                validade_dt = datetime.strptime(validade, '%Y-%m-%d')
                validade_int = int(validade_dt.timestamp())
                
                # Buscar produtos com este código de barras e data de validade
                produto = cursor.execute('''
                    SELECT * FROM estoque 
                    WHERE codigo_de_barras = ? AND validade_int = ?
                    AND quantidade > 0
                ''', (codigo_de_barras, validade_int)).fetchone()
                
                if produto:
                    return redirect(url_for('retirada_com_id', produto_id=produto['id']))
                else:
                    error = translate('product_not_found', lang)
                    return render_template('retirar.html', error=error)
            except ValueError:
                error = translate('invalid_date', lang)
                return render_template('retirar.html', error=error)
            # Busca o produto exato no estoque (caso use lote)
            produto = cursor.execute('SELECT * FROM estoque WHERE codigo_de_barras = ? AND lote = ?', 
                                     (codigo_de_barras, lote)).fetchone()
            
            if produto:
                id_encontrado = produto['id']
                # Redireciona para a tela de confirmação (retirada_com_id)
                return redirect(url_for('retirada_com_id', produto_id=id_encontrado))
            else:
                flash(translate('product_not_found', lang), 'error')
                return render_template('retirar.html') 
                
    return render_template('retirar.html')


@app.route('/api/produtos_por_codigo')
def api_produtos_por_codigo():
    """Retorna JSON com todas as entradas no estoque para um código de barras.

    Query params:
      - codigo: código de barras (string)
    Retorna lista de objetos: id, produto_nome, validade_text, validade_int, quantidade, lote, image_path
    """
    codigo = request.args.get('codigo', '').strip()
    if not codigo:
        return jsonify([])

    db = get_db()
    cursor = db.cursor()
    try:
        rows = cursor.execute('''
            SELECT id, produto_nome, validade_text, validade_int, quantidade, lote, image_path
            FROM estoque
            WHERE codigo_de_barras = ?
            ORDER BY validade_int ASC
        ''', (codigo,)).fetchall()

        resultados = []
        for r in rows:
            resultados.append({
                'id': r['id'],
                'produto_nome': r['produto_nome'],
                'validade_text': r['validade_text'],
                'validade_int': r['validade_int'],
                'quantidade': r['quantidade'],
                'lote': r['lote'],
                'image_path': r['image_path']
            })

        return jsonify(resultados)
    except Exception as e:
        # Em caso de erro, retorna lista vazia para o front-end tratar
        print('Erro ao buscar produtos por codigo:', e)
        return jsonify([]), 500


@app.route('/retirada_estoque/<int:produto_id>', methods=['GET', 'POST'])
def retirada_com_id(produto_id):
    """Exibe o produto e gerencia a submissão da quantidade a ser retirada."""
    db = get_db()
    cursor = db.cursor()
    
    # 1. Busca o produto para exibição/validação
    produto = cursor.execute('SELECT * FROM estoque WHERE id = ?', (produto_id,)).fetchone()
    
    lang = session.get('lang', 'pt')
    if request.method == 'POST':
        try:
            # Garante que a quantidade seja um número inteiro válido
            quantidade_retirada = int(request.form['quantidade'])
            responsavel = request.form.get('responsavel', '').strip()
        except (KeyError, ValueError):
            error = translate('quantity_must_be_positive', lang)
            return render_template('retirar_estoque.html', produto=produto, produto_id=produto_id, error=error)
            
        
        # 2. Validação: Verifica se o estoque é suficiente
        if produto and quantidade_retirada > 0 and produto['quantidade'] >= quantidade_retirada:
            
            nova_quantidade = produto['quantidade'] - quantidade_retirada
            
            # 3. ATUALIZA ESTOQUE
            cursor.execute('UPDATE estoque SET quantidade = ? WHERE id = ?', (nova_quantidade, produto_id))
            
            # 4. REGISTRA MOVIMENTAÇÃO COM RESPONSÁVEL (armazenado em motivo para compatibilidade)
            cursor.execute('''
                INSERT INTO movimentacao (product_id, product_barcode, name, action, quantidade, motivo)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (produto['id'], produto['codigo_de_barras'], produto['produto_nome'], 'retirada', quantidade_retirada, responsavel))
            
            # 5. EXECUTA A LIMPEZA AUTOMÁTICA
            # Remove itens que ficaram com quantidade <= 0 após a retirada (o commit finaliza tudo)
            limpar_estoque_zerado()
            
            flash(translate('success', lang) + f" - {quantidade_retirada} {translate('units', lang)}", 'success')
            return redirect(url_for('retirada'))
        
        else:
            error = translate('insufficient_stock', lang)
            return render_template('retirar_estoque.html', produto=produto, produto_id=produto_id, error=error)

    # Exibe a página de retirada (GET)
    return render_template('retirar_estoque.html', produto=produto, produto_id=produto_id)


# Endpoint AJAX para ajuste rápido de quantidade (incremento/decremento)
@app.route('/api/adjust_quantity', methods=['POST'])
def api_adjust_quantity():
    db = get_db()
    cursor = db.cursor()
    try:
        data = request.get_json()
        product_id = int(data.get('product_id'))
        action = data.get('action')

        produto = cursor.execute('SELECT * FROM estoque WHERE id = ?', (product_id,)).fetchone()
        if not produto:
            return {'ok': False, 'error': 'Produto não encontrado'}, 404

        if action == 'add':
            novo = produto['quantidade'] + 1
            movimento = 'entrada'
            cursor.execute('UPDATE estoque SET quantidade = ? WHERE id = ?', (novo, product_id))
            cursor.execute('INSERT INTO movimentacao (product_id, product_barcode, name, action, quantidade, motivo) VALUES (?, ?, ?, ?, ?, ?)',
                       (product_id, produto['codigo_de_barras'], produto['produto_nome'], movimento, 1, 'Adição rápida'))
        elif action == 'remove':
            # Remove não faz mais nada - modal cuida disso
            return {'ok': False, 'error': 'Use o modal para retiradas'}, 400
        else:
            return {'ok': False, 'error': 'Ação inválida'}, 400

        if novo < 0:
            return {'ok': False, 'error': 'Quantidade não pode ser negativa'}, 400

        db.commit()

        # Limpa itens zerados
        limpar_estoque_zerado()

        return {'ok': True, 'new_quantity': novo}
    except Exception as e:
        db.rollback()
        return {'ok': False, 'error': str(e)}, 500


# Nova rota para retirada com motivo (via modal)
@app.route('/api/retirar_com_motivo', methods=['POST'])
def api_retirar_com_motivo():
    db = get_db()
    cursor = db.cursor()
    try:
        data = request.get_json()
        product_id = int(data.get('product_id'))
        quantidade = int(data.get('quantidade', 1))
        motivo = data.get('motivo', '').strip()

        if not motivo:
            return {'ok': False, 'error': 'Motivo é obrigatório'}, 400

        produto = cursor.execute('SELECT * FROM estoque WHERE id = ?', (product_id,)).fetchone()
        if not produto:
            return {'ok': False, 'error': 'Produto não encontrado'}, 404

        if quantidade > produto['quantidade']:
            return {'ok': False, 'error': 'Quantidade insuficiente em estoque'}, 400

        novo = produto['quantidade'] - quantidade
        
        cursor.execute('UPDATE estoque SET quantidade = ? WHERE id = ?', (novo, product_id))
        cursor.execute('''INSERT INTO movimentacao 
            (product_id, product_barcode, name, action, quantidade, motivo) 
            VALUES (?, ?, ?, ?, ?, ?)''',
            (product_id, produto['codigo_de_barras'], produto['produto_nome'], 
             'retirada', quantidade, motivo))
        db.commit()

        # Limpa itens zerados
        limpar_estoque_zerado()

        return {'ok': True, 'new_quantity': novo}
    except Exception as e:
        db.rollback()
        return {'ok': False, 'error': str(e)}, 500


# Nova rota para adição com motivo (via modal)
@app.route('/api/adicionar_com_motivo', methods=['POST'])
def api_adicionar_com_motivo():
    db = get_db()
    cursor = db.cursor()
    try:
        data = request.get_json()
        product_id = int(data.get('product_id'))
        quantidade = int(data.get('quantidade', 1))
        motivo = data.get('motivo', '').strip()

        if not motivo:
            return {'ok': False, 'error': 'Motivo é obrigatório'}, 400

        produto = cursor.execute('SELECT * FROM estoque WHERE id = ?', (product_id,)).fetchone()
        if not produto:
            return {'ok': False, 'error': 'Produto não encontrado'}, 404

        novo = produto['quantidade'] + quantidade
        
        cursor.execute('UPDATE estoque SET quantidade = ? WHERE id = ?', (novo, product_id))
        cursor.execute('''INSERT INTO movimentacao 
            (product_id, product_barcode, name, action, quantidade, motivo) 
            VALUES (?, ?, ?, ?, ?, ?)''',
            (product_id, produto['codigo_de_barras'], produto['produto_nome'], 
             'entrada', quantidade, motivo))
        db.commit()

        return {'ok': True, 'new_quantity': novo}
    except Exception as e:
        db.rollback()
        return {'ok': False, 'error': str(e)}, 500


@app.route('/historico')
@login_required
def historico():
    """Exibe o histórico de movimentações com filtros."""
    from datetime import timedelta
    db = get_db()
    cursor = db.cursor()

    # Parâmetros de query (filtros e paginação)
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    periodo = request.args.get('periodo', 'today')  # Novo parâmetro de período
    custom_date = request.args.get('date', '')  # Data customizada (quando periodo='custom')
    action_filter = request.args.get('action')
    busca = request.args.get('q', '').strip()

    # Calcula as datas baseado no período selecionado
    hoje = date.today()
    date_from = None
    date_to = None
    periodo_display = ''
    
    if periodo == 'today':
        date_from = date_to = hoje
        periodo_display = f"{hoje.strftime('%d/%m/%Y')}"
    elif periodo == 'week':
        date_from = hoje - timedelta(days=7)
        date_to = hoje
        periodo_display = f"{date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}"
    elif periodo == 'month':
        date_from = hoje - timedelta(days=30)
        date_to = hoje
        periodo_display = f"{date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}"
    elif periodo == 'current_month':
        date_from = date(hoje.year, hoje.month, 1)
        date_to = hoje
        periodo_display = f"{date_from.strftime('%B/%Y')}"
    elif periodo == 'last_month':
        # Primeiro dia do mês passado
        if hoje.month == 1:
            date_from = date(hoje.year - 1, 12, 1)
            ultimo_dia = date(hoje.year, 1, 1) - timedelta(days=1)
        else:
            date_from = date(hoje.year, hoje.month - 1, 1)
            ultimo_dia = date(hoje.year, hoje.month, 1) - timedelta(days=1)
        date_to = ultimo_dia
        periodo_display = f"{date_from.strftime('%B/%Y')}"
    elif periodo == 'year':
        date_from = date(hoje.year, 1, 1)
        date_to = hoje
        periodo_display = f"{hoje.year}"
    elif periodo == 'custom' and custom_date:
        try:
            date_from = date_to = date.fromisoformat(custom_date)
            periodo_display = f"{date_from.strftime('%d/%m/%Y')}"
        except ValueError:
            date_from = date_to = hoje
            periodo_display = f"{hoje.strftime('%d/%m/%Y')}"
    elif periodo == 'all':
        # Sem filtro de data - mostra tudo
        date_from = date_to = None
        periodo_display = "Todas as movimentações"
    else:
        # Padrão: hoje
        date_from = date_to = hoje
        periodo_display = f"{hoje.strftime('%d/%m/%Y')}"

    # Monta cláusula WHERE dinâmica
    where_clauses = []
    params = []

    if action_filter:
        where_clauses.append('action = ?')
        params.append(action_filter)

    if busca:
        where_clauses.append('(LOWER(name) LIKE ? OR product_barcode LIKE ?)')
        search_term = f"%{busca.lower()}%"
        params.extend([search_term, search_term])

    # Aplica filtros de data apenas se não for 'all'
    if date_from and date_to:
        if date_from == date_to:
            where_clauses.append('DATE(timestamp) = DATE(?)')
            params.append(date_from.isoformat())
        else:
            where_clauses.append('DATE(timestamp) >= DATE(?)')
            params.append(date_from.isoformat())
            where_clauses.append('DATE(timestamp) <= DATE(?)')
            params.append(date_to.isoformat())

    where_sql = ''
    if where_clauses:
        where_sql = 'WHERE ' + ' AND '.join(where_clauses)

    # Conta total de items (para paginação)
    count_sql = f'SELECT COUNT(*) as cnt FROM movimentacao {where_sql}'
    total_rows = cursor.execute(count_sql, params).fetchone()['cnt']

    total_pages = max(1, math.ceil(total_rows / per_page))
    if page < 1:
        page = 1
    if page > total_pages:
        page = total_pages

    offset = (page - 1) * per_page

    # Busca as movimentações com paginação
    sql = f'''SELECT *, datetime(timestamp) as timestamp 
             FROM movimentacao {where_sql} 
             ORDER BY timestamp DESC 
             LIMIT ? OFFSET ?'''
    final_params = params + [per_page, offset]
    movimentacoes_raw = cursor.execute(sql, final_params).fetchall()

    # Busca datas disponíveis para o seletor
    datas_disponiveis = cursor.execute('''
        SELECT DISTINCT DATE(timestamp) as data 
        FROM movimentacao 
        ORDER BY data DESC
    ''').fetchall()
    datas_disponiveis = [row['data'] for row in datas_disponiveis]

    # Normaliza o campo timestamp para objetos datetime (o template usa .strftime)
    from datetime import datetime as _dt
    from types import SimpleNamespace

    movimentacoes = []
    for mov in movimentacoes_raw:
        # mov pode ser sqlite3.Row; converte para dict primeiro
        try:
            row = dict(mov)
        except Exception:
            # se já for dict
            row = mov

        ts = row.get('timestamp')
        parsed = ts
        if isinstance(ts, (int, float)):
            try:
                parsed = _dt.fromtimestamp(ts)
            except Exception:
                parsed = ts
        elif isinstance(ts, str):
            # tenta vários formatos comuns
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d'):
                try:
                    parsed = _dt.strptime(ts, fmt)
                    break
                except Exception:
                    try:
                        parsed = _dt.fromisoformat(ts)
                        break
                    except Exception:
                        parsed = ts

        row['timestamp'] = parsed
        # transforma em um objeto com atributos para facilitar o template
        movimentacoes.append(SimpleNamespace(**row))

    return render_template('historico.html',
                         movimentacoes=movimentacoes,
                         current_page=page,
                         total_pages=total_pages,
                         has_prev=page > 1,
                         has_next=page < total_pages,
                         periodo=periodo,
                         periodo_display=periodo_display,
                         custom_date=custom_date or hoje.isoformat(),
                         total_movimentacoes=total_rows,
                         today_date=date.today().isoformat())

@app.route('/exportar_historico')
@login_required
def exportar_historico():
    """Exporta o histórico de movimentações do período selecionado para Excel (.xlsx)."""
    from datetime import timedelta
    db = get_db()
    cursor = db.cursor()
    
    # Parâmetros
    periodo = request.args.get('periodo', 'today')
    custom_date = request.args.get('date', '')
    
    # Calcula as datas baseado no período selecionado (mesma lógica da visualização)
    hoje = date.today()
    date_from = None
    date_to = None
    periodo_nome = ''
    
    if periodo == 'today':
        date_from = date_to = hoje
        periodo_nome = f"{hoje.strftime('%d-%m-%Y')}"
    elif periodo == 'week':
        date_from = hoje - timedelta(days=7)
        date_to = hoje
        periodo_nome = f"Ultimos_7_dias_{date_to.strftime('%d-%m-%Y')}"
    elif periodo == 'month':
        date_from = hoje - timedelta(days=30)
        date_to = hoje
        periodo_nome = f"Ultimos_30_dias_{date_to.strftime('%d-%m-%Y')}"
    elif periodo == 'current_month':
        date_from = date(hoje.year, hoje.month, 1)
        date_to = hoje
        periodo_nome = f"{date_from.strftime('%B_%Y')}"
    elif periodo == 'last_month':
        if hoje.month == 1:
            date_from = date(hoje.year - 1, 12, 1)
            ultimo_dia = date(hoje.year, 1, 1) - timedelta(days=1)
        else:
            date_from = date(hoje.year, hoje.month - 1, 1)
            ultimo_dia = date(hoje.year, hoje.month, 1) - timedelta(days=1)
        date_to = ultimo_dia
        periodo_nome = f"{date_from.strftime('%B_%Y')}"
    elif periodo == 'year':
        date_from = date(hoje.year, 1, 1)
        date_to = hoje
        periodo_nome = f"Ano_{hoje.year}"
    elif periodo == 'custom' and custom_date:
        try:
            date_from = date_to = date.fromisoformat(custom_date)
            periodo_nome = f"{date_from.strftime('%d-%m-%Y')}"
        except ValueError:
            date_from = date_to = hoje
            periodo_nome = f"{hoje.strftime('%d-%m-%Y')}"
    elif periodo == 'all':
        date_from = date_to = None
        periodo_nome = "Todas_movimentacoes"
    else:
        date_from = date_to = hoje
        periodo_nome = f"{hoje.strftime('%d-%m-%Y')}"
    
    # Monta a query baseado no período
    if date_from and date_to:
        if date_from == date_to:
            where_clause = 'WHERE DATE(timestamp) = DATE(?)'
            params = (date_from.isoformat(),)
        else:
            where_clause = 'WHERE DATE(timestamp) >= DATE(?) AND DATE(timestamp) <= DATE(?)'
            params = (date_from.isoformat(), date_to.isoformat())
    else:
        # Todas as movimentações
        where_clause = ''
        params = ()
    
    # Busca as movimentações
    movimentacoes = cursor.execute(f'''
        SELECT 
            datetime(timestamp) as timestamp,
            product_barcode,
            name,
            action,
            quantidade,
            motivo
        FROM movimentacao 
        {where_clause}
        ORDER BY timestamp DESC
    ''', params).fetchall()
    
    # Cria o workbook do Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Histórico {periodo_nome[:25]}"  # Limite de 31 caracteres para título de aba
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4D7033", end_color="4D7033", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Cabeçalhos
    headers = ['Data/Hora', 'Código de Barras', 'Produto', 'Ação', 'Quantidade', 'Responsável']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for row_num, mov in enumerate(movimentacoes, 2):
        # Formata timestamp
        try:
            if isinstance(mov['timestamp'], str):
                ts = datetime.fromisoformat(mov['timestamp'].replace(' ', 'T'))
            else:
                ts = mov['timestamp']
            timestamp_str = ts.strftime('%d/%m/%Y %H:%M:%S')
        except:
            timestamp_str = str(mov['timestamp'])
        
        ws.cell(row=row_num, column=1).value = timestamp_str
        ws.cell(row=row_num, column=2).value = mov['product_barcode']
        ws.cell(row=row_num, column=3).value = mov['name']
        
        # Ação com cor
        action_cell = ws.cell(row=row_num, column=4)
        action_cell.value = mov['action'].title()
        if mov['action'] == 'entrada':
            action_cell.font = Font(color="006400", bold=True)  # Verde escuro
        else:
            action_cell.font = Font(color="8B0000", bold=True)  # Vermelho escuro
        
        ws.cell(row=row_num, column=5).value = mov['quantidade']
        ws.cell(row=row_num, column=6).value = mov['motivo'] or '-'
    
    # Ajusta largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40
    
    # Salva em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    filename = f"historico_{periodo_nome}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))

# --------------------------------------------------------------------------
# API DE RESPONSÁVEIS
# --------------------------------------------------------------------------

@app.route('/api/responsaveis', methods=['GET'])
@login_required
def get_responsaveis():
    """Retorna a lista de responsáveis cadastrados."""
    db = get_db()
    cursor = db.cursor()
    
    cursor.execute('SELECT id, nome FROM responsaveis ORDER BY nome ASC')
    responsaveis = cursor.fetchall()
    
    return jsonify({
        'ok': True,
        'responsaveis': [{'id': r['id'], 'nome': r['nome']} for r in responsaveis]
    })

@app.route('/api/responsaveis', methods=['POST'])
@login_required
def add_responsavel():
    """Adiciona um novo responsável."""
    data = request.get_json()
    nome = data.get('nome', '').strip()
    
    if not nome:
        return jsonify({'ok': False, 'error': 'Nome é obrigatório'}), 400
    
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute('INSERT INTO responsaveis (nome) VALUES (?)', (nome,))
        db.commit()
        return jsonify({'ok': True, 'id': cursor.lastrowid})
    except sqlite3.IntegrityError:
        return jsonify({'ok': False, 'error': 'Nome já cadastrado'}), 400
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

@app.route('/api/responsaveis/<int:responsavel_id>', methods=['DELETE'])
@login_required
def delete_responsavel(responsavel_id):
    """Remove um responsável."""
    db = get_db()
    cursor = db.cursor()
    
    try:
        cursor.execute('DELETE FROM responsaveis WHERE id = ?', (responsavel_id,))
        db.commit()
        
        if cursor.rowcount == 0:
            return jsonify({'ok': False, 'error': 'Responsável não encontrado'}), 404
            
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500

# --------------------------------------------------------------------------
# INICIALIZAÇÃO DA APLICAÇÃO
# --------------------------------------------------------------------------

@app.cli.command('init-db')
def init_db_command():
    """Comando para inicializar o banco de dados (ex: flask init-db)."""
    with app.app_context():
        init_db()
        print('Banco de dados inicializado.')

# Rota de debug (apenas para ambiente de desenvolvimento) para listar templates e static
@app.route('/debug/list_files')
@login_required
def debug_list_files():
    """Lista arquivos em templates/ e static/ — útil para detectar nomes/path incorretos."""
    def list_dir_tree(base):
        result = []
        for root, dirs, files in os.walk(base):
            for f in files:
                rel = os.path.relpath(os.path.join(root, f), app.root_path)
                result.append(rel)
        return sorted(result)
    templates = list_dir_tree(os.path.join(app.root_path, 'templates')) if os.path.isdir(os.path.join(app.root_path, 'templates')) else []
    static_files = list_dir_tree(app.static_folder) if os.path.isdir(app.static_folder) else []
    return jsonify({'templates': templates, 'static': static_files})

@app.route('/verify_resources')
@login_required
def verify_resources():
    """Verifica integridade de recursos estáticos e templates."""
    problems = []
    
    # Verifica templates
    templates_dir = app.template_folder
    if os.path.exists(templates_dir):
        for root, dirs, files in os.walk(templates_dir):
            for file in files:
                if file.endswith('.html'):
                    filepath = os.path.join(root, file)
                    try:
                        with open(filepath, 'r', encoding='utf-8') as f:
                            content = f.read()
                            # Verifica links CSS
                            import re
                            css_links = re.findall(r'href=["\']([^"\']*\.css)["\']', content)
                            for link in css_links:
                                if link.startswith('/static/'):
                                    css_path = os.path.join(app.root_path, link[1:])
                                    if not os.path.exists(css_path):
                                        problems.append(f"CSS não encontrado: {link} (referenciado em {file})")
                            
                            # Verifica links JS
                            js_links = re.findall(r'src=["\']([^"\']*\.js)["\']', content)
                            for link in js_links:
                                if link.startswith('/static/'):
                                    js_path = os.path.join(app.root_path, link[1:])
                                    if not os.path.exists(js_path):
                                        problems.append(f"JS não encontrado: {link} (referenciado em {file})")
                            
                            # Verifica imagens
                            img_links = re.findall(r'src=["\']([^"\']*\.(png|jpg|jpeg|gif|webp|svg))["\']', content, re.IGNORECASE)
                            for link, ext in img_links:
                                if link.startswith('/static/'):
                                    img_path = os.path.join(app.root_path, link[1:])
                                    if not os.path.exists(img_path):
                                        problems.append(f"Imagem não encontrada: {link} (referenciado em {file})")
                    except Exception as e:
                        problems.append(f"Erro ao processar {file}: {str(e)}")
    
    # Verifica arquivos estáticos órfãos
    static_files = []
    if os.path.exists(app.static_folder):
        for root, dirs, files in os.walk(app.static_folder):
            for file in files:
                rel_path = os.path.relpath(os.path.join(root, file), app.static_folder)
                static_files.append(rel_path.replace('\\', '/'))
    
    return jsonify({
        'ok': len(problems) == 0,
        'problems': problems,
        'static_files_count': len(static_files),
        'static_files': static_files[:50]  # Primeiros 50 para não sobrecarregar
    })

# Handler 404 personalizado
@app.errorhandler(404)
def not_found_error(error):
    """Handler para páginas não encontradas."""
    app.logger.warning(f"404 Error: {request.url} - {error.description}")
    return render_template('error.html', 
                         error="Página não encontrada", 
                         error_code=404,
                         error_description=error.description), 404

# Handler 500 para log melhor (dev)
@app.errorhandler(500)
def internal_error(error):
    app.logger.error(f"Internal Server Error: {error}", exc_info=True)
    # tenta renderizar template de erro se existir
    try:
        return render_template('error.html', error=error), 500
    except Exception:
        return "Erro interno do servidor. Verifique os logs.", 500

if __name__ == '__main__':
    with app.app_context():
        # Inicializa o banco de dados antes de iniciar o servidor
        init_db()
        
    # Coloque o modo debug como False em produção
    app.run(debug=True)
